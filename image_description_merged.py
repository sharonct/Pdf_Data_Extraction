from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

def fill_missing_images(text_excel, images_excel, output_excel):
    # Load the text Excel file
    wb_text = load_workbook(text_excel)
    ws_text = wb_text.active

    # Load the images Excel file
    wb_images = load_workbook(images_excel)
    ws_images = wb_images.active

    # Create a new workbook for the output
    wb_output = load_workbook(text_excel)
    ws_output = wb_output.active

    # Dictionary to store trademark numbers and corresponding image paths
    image_paths = {}

    # Extract trademark numbers and image paths from the images Excel file
    for row in ws_images.iter_rows(min_row=2, max_col=3, values_only=True):
        trademark_number = row[0]
        image_path = row[2]
        image_paths[trademark_number] = image_path

    # Check for null image entries in the text Excel file and fill them with images from the images Excel file
    row_idx = 2  # Start from the second row
    for row in ws_text.iter_rows(min_row=2, max_col=6, values_only=True):
        trademark_number = row[0]
        image_path = row[5]

        if image_path is None:
            # If image is missing, try to fill it from the image_paths dictionary
            if trademark_number in image_paths:
                new_image_path = image_paths[trademark_number]
                if new_image_path is not None:
                    # Add the image to the output worksheet
                    img_xl = XLImage(new_image_path)
                    img_xl.anchor = ws_output.cell(row=row_idx, column=6).coordinate
                    ws_output.add_image(img_xl)

        row_idx += 1

    # Save the output Excel file
    wb_output.save(output_excel)