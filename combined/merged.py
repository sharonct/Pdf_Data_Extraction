from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
import pandas as pd

def fill_missing_images(text_df, images_df, output_excel):
    # Create a new workbook for the output
    wb_output = Workbook()
    ws_output = wb_output.active

    # Add headers from the text dataframe
    for col_idx, col_name in enumerate(text_df.columns, start=1):
        ws_output.cell(row=1, column=col_idx, value=col_name)

    # Dictionary to store trademark numbers and corresponding images
    image_data = {}

    # Extract trademark numbers and images from the images dataframe
    for index, row in images_df.iterrows():
        trademark_number = row['TrademarkNo']
        image_bytes = row['ImageData']
        if trademark_number and image_bytes:
            image_data[trademark_number] = image_bytes

    # Iterate over rows in the text dataframe
    for index, row in text_df.iterrows():
        trademark_number = row['Trademark Number (210)']
        image_path = row['Image/Mark']

        # If image is missing or empty, try to fill it from the image_data dictionary
        if pd.isnull(image_path) or image_path == "":
            if trademark_number in image_data:
                image_bytes = image_data[trademark_number]
                if image_bytes:
                    img_buffer = BytesIO(image_bytes)
                    pil_image = PILImage.open(img_buffer)
                    img_xl = XLImage(img_buffer)
                    cell = ws_output.cell(row=index+2, column=text_df.columns.get_loc('Image/Mark') + 1)  # Assuming Image/Mark is in the dataframe
                    ws_output.add_image(img_xl, cell.coordinate)

        # Copy the existing row from text dataframe
        for col_idx, value in enumerate(row, start=1):
            if pd.notnull(value):
                ws_output.cell(row=index+2, column=col_idx, value=value)

    # Save the output Excel file
    wb_output.save(output_excel)

