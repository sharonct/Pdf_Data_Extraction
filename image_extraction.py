import os  # Import the os module for file operations
from openpyxl import Workbook  # Import Workbook class from openpyxl for Excel operations
from openpyxl.drawing.image import Image as XLImage  # Import Image class from openpyxl for handling images in Excel
import fitz  # Import fitz module from PyMuPDF for PDF operations
from PIL import Image as PILImage  # Import Image class from PIL for image operations
import re  # Import re module for regular expression operations

def extract_trademarks_and_logos(pdf_file, output_excel, start_page, end_page):
    # Define paths for the folders
    images_folder = "Images"
    data_folder = "Data"
    images_path = os.path.join(data_folder, images_folder)  # Path for the images folder
    excel_path = os.path.join(data_folder, output_excel) + ".xlsx"  # Path for the Excel file

    # Clear existing images in the folder
    for file in os.listdir(images_path):
        os.remove(os.path.join(images_path, file))

    # Create the folders if they don't exist
    os.makedirs(images_path, exist_ok=True)

    # Create a workbook and add a worksheet
    wb = Workbook()  # Create a new Workbook object
    ws = wb.active  # Get the active worksheet in the workbook
    ws.append(['TradeMarkNo.', 'Images'])  # Add headers to the worksheet

    doc = fitz.open(pdf_file)  # Open the PDF file
    row = 2  # Start from the second row to leave space for the header

    for page_num, page in enumerate(doc, start=1):  # Iterate through pages with page number starting from 1
        if start_page <= page_num <= end_page:  # Check if the page is within the specified range
            text = page.get_text()  # Extract text from the page
            images = page.get_images(full=True)  # Extract images from the page

            # Extract trademark numbers using regular expressions
            trademark_numbers = re.findall(r'\(210\): (\d+) \(220\)', text)

            logos = []  # List to store image paths
            image_index = 0  # Index for iterating through images

            # Iterate through trademark numbers found on the page
            for trademark_number in trademark_numbers:
                last_word_caps = trademark_number.split()[-1].isupper()  # Check if the last word is in uppercase

                if not last_word_caps and image_index < len(images):  # If last word is not in uppercase and images exist
                    try:
                        # Save image with trademark number as filename in the images folder
                        image_path = os.path.join(images_path, f"{trademark_number}.png")
                        pix = fitz.Pixmap(doc, images[image_index][0])
                        pil_image = PILImage.frombytes("RGB", [pix.width, pix.height], pix.samples)
                        pil_image = pil_image.resize((50, 50))  # Resize the image to 50x50 pixels
                        pil_image.save(image_path)  # Save the image
                        logos.append(image_path)  # Add the image path to the logos list
                        image_index += 1  # Increment image index
                    except Exception as e:
                        print(f"Error saving image: {e}")
                else:
                    logos.append(None)  # Append None if no image found for the trademark number

            # Iterate through trademark numbers and corresponding images
            for idx, trademark_number in enumerate(trademark_numbers):
                image_path = logos[idx]  # Get the image path corresponding to the trademark number
                if image_path:
                    img_xl = XLImage(image_path)  # Create an XLImage object with the image path
                    img_xl.anchor = ws.cell(row=row, column=2).coordinate  # Set the anchor point for the image
                    ws.add_image(img_xl)  # Add the image to the worksheet
                else:
                    ws.cell(row=row, column=2).value = "No Image"  # If no image, set cell value to "No Image"
                ws.cell(row=row, column=1).value = trademark_number  # Set the trademark number in the worksheet
                row += 1  # Move to the next row for the next trademark number

    wb.save(excel_path)  # Save the workbook to the specified Excel file
    print("Extraction completed.")  # Print message indicating extraction completion

if __name__ == "__main__":
    # Prompt user to enter PDF file name, Excel file name, and page range for extraction
    pdf_file = input("Enter the name of the PDF file: ")
    output_excel = input("Enter the name of the Excel file: ")
    start_page = int(input("Enter the start page number: "))
    end_page = int(input("Enter the end page number: "))

    # Call the extraction function with provided inputs
    extract_trademarks_and_logos(pdf_file, output_excel, start_page, end_page)
